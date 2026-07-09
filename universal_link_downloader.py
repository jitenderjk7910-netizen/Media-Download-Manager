import base64
import csv
import hashlib
import html
import importlib.util
import mimetypes
import os
import queue
import re
import shutil
import subprocess
import sys
import tempfile
import threading
import time
import urllib.parse
import zipfile
from concurrent.futures import FIRST_COMPLETED, ThreadPoolExecutor, wait
from dataclasses import dataclass
from pathlib import Path
try:
    from tkinter import filedialog, messagebox, ttk
    from tkinter.scrolledtext import ScrolledText
    import tkinter as tk
    TKINTER_AVAILABLE = True
except (ImportError, Exception):
    tk = None
    ttk = None
    filedialog = None
    messagebox = None
    ScrolledText = None
    TKINTER_AVAILABLE = False

from typing import Callable, Dict, List, Optional, Tuple

import requests

try:
    import openpyxl
except ImportError:
    openpyxl = None

try:
    import gdown
except ImportError:
    gdown = None

# Centralized configuration & named constants (see app_config.py).
from app_config import (
    APP_VERSION as _APP_VERSION,
    CHUNK_SIZE,
    DEFAULT_TIMEOUT,
    DEFAULT_USER_AGENT,
    DEFAULT_WORKERS,
    DROPBOX_USER_AGENT,
    LOG_NAME,
    MAGIC_BYTE_BUFFER,
    MAX_DISPLAYED_ROWS,
    MAX_ONEDRIVE_IMAGES,
    MAX_PAGE_IMAGES,
    MAX_RETRIES_LIMIT,
    MAX_WORKERS,
    MIN_TIMEOUT,
    MIN_WORKERS,
    PAUSE_POLL_INTERVAL,
    PROVIDER_LIMITS,
    RETRY_BACKOFF_MAX,
    RETRYABLE_MESSAGE_FRAGMENTS,
    RETRYABLE_STATUSES,
    THUMB_EXTENSIONS,
    VALIDATION_BUFFER,
    IMAGE_EXTENSIONS as _IMAGE_EXTS,
    VIDEO_EXTENSIONS as _VIDEO_EXTS,
    ARCHIVE_EXTENSIONS as _ARCHIVE_EXTS,
)

APP_TITLE = "Media Download Manager"
APP_VERSION = _APP_VERSION
# Re-exported for existing imports (e.g. web_app.py imports IMAGE_EXTENSIONS).
IMAGE_EXTENSIONS = _IMAGE_EXTS
VIDEO_EXTENSIONS = _VIDEO_EXTS
ARCHIVE_EXTENSIONS = _ARCHIVE_EXTS


@dataclass
class LinkTask:
    row: int
    folder: str
    url: str
    row_data: Optional[dict] = None


@dataclass
class DownloadResult:
    status: str
    files: int
    message: str
    bytes_downloaded: int = 0
    attempts: int = 1
    duration: float = 0.0


def sanitize_name(value: str) -> str:
    value = str(value or "").strip()
    value = re.sub(r'[<>:"/\\|?*\x00-\x1f]', "_", value)
    value = value.strip(" .")
    return value or "unknown"


def is_url(value: str) -> bool:
    return str(value or "").strip().lower().startswith(("http://", "https://"))


def open_in_file_manager(path) -> None:
    """Open a path in the OS file manager (cross-platform).

    Replaces the Windows-only ``os.startfile`` call so the tool also works on
    macOS and Linux.
    """
    target = Path(path) if path else None
    if not target:
        return
    resolved = str(target.resolve())
    if sys.platform.startswith("win"):
        os.startfile(resolved)  # type: ignore[attr-defined]
    elif sys.platform == "darwin":
        subprocess.Popen(["open", resolved])
    else:
        subprocess.Popen(["xdg-open", resolved])


def unique_path(folder, filename):
    folder = Path(folder)
    folder.mkdir(parents=True, exist_ok=True)
    filename = sanitize_name(filename)
    stem = sanitize_name(Path(filename).stem)
    suffix = Path(filename).suffix
    candidate = folder / f"{stem}{suffix}"
    count = 2
    while candidate.exists():
        candidate = folder / f"{stem}_{count}{suffix}"
        count += 1
    return candidate


def filename_from_url(url, content_type=""):
    parsed = urllib.parse.urlsplit(url)
    name = Path(urllib.parse.unquote(parsed.path)).name
    if not name or "." not in name:
        ext = mimetypes.guess_extension((content_type or "").split(";")[0].strip()) or ".bin"
        name = f"download{ext}"
    return sanitize_name(name)


def load_sharepoint_core():
    here = Path(__file__).resolve()
    root = here.parents[1]
    core_path = root / "1Sharepoint & Onedrive Folder Downloader" / "sharepoint_downloader.py"
    if not core_path.exists():
        return None
    spec = importlib.util.spec_from_file_location("sharepoint_downloader_core", core_path)
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


SHAREPOINT_CORE = load_sharepoint_core()


def provider_name(url):
    host = urllib.parse.urlsplit(url).netloc.lower()
    if "dropbox.com" in host:
        return "dropbox"
    if "drive.google.com" in host or "docs.google.com" in host:
        return "google_drive"
    if "sharepoint.com" in host or "onedrive.live.com" in host or "1drv.ms" in host:
        return "sharepoint_onedrive"
    if "ibb.co" in host or "imgbb.com" in host:
        return "imgbb"
    return "generic"


def normalize_dropbox(url):
    parts = urllib.parse.urlsplit(url)
    query = dict(urllib.parse.parse_qsl(parts.query, keep_blank_values=True))
    query["dl"] = "1"
    query.pop("raw", None)
    return urllib.parse.urlunsplit(
        (parts.scheme, parts.netloc, parts.path, urllib.parse.urlencode(query), parts.fragment)
    )


def is_dropbox_folder(url):
    path = urllib.parse.urlsplit(url).path.lower()
    return "/scl/fo/" in path or "/sh/" in path


def is_sharepoint_folder_share(url):
    path = urllib.parse.urlsplit(url).path.lower()
    return "/:f:/" in path or "onedrive.aspx" in path and "id=" in urllib.parse.urlsplit(url).query.lower()


def google_drive_id(url):
    patterns = [
        r"/file/d/([a-zA-Z0-9_-]+)",
        r"/folders/([a-zA-Z0-9_-]+)",
        r"[?&]id=([a-zA-Z0-9_-]+)",
    ]
    for pattern in patterns:
        match = re.search(pattern, url)
        if match:
            return match.group(1)
    return None


def google_drive_is_folder(url):
    return "/folders/" in url or "drive/folders" in url


def is_probably_image_url(url):
    path = urllib.parse.urlsplit(url).path.lower()
    return Path(path).suffix in IMAGE_EXTENSIONS


def is_internal_metadata_file(path_or_name) -> bool:
    path = Path(str(path_or_name or ""))
    name = path.name.lower()
    if name in {"download.json", "metadata.json", ".download.json"}:
        return True
    if name.startswith("download") and path.suffix.lower() == ".json":
        return True
    return False


def image_links_from_html(text, base_url):
    text = html.unescape(text or "")
    text = text.replace("\\/", "/")
    text = text.replace("\\u0026", "&")
    candidates = []

    meta_patterns = [
        r'<meta[^>]+property=["\']og:image["\'][^>]+content=["\']([^"\']+)["\']',
        r'<meta[^>]+content=["\']([^"\']+)["\'][^>]+property=["\']og:image["\']',
        r'<link[^>]+rel=["\']image_src["\'][^>]+href=["\']([^"\']+)["\']',
    ]
    for pattern in meta_patterns:
        candidates.extend(re.findall(pattern, text, flags=re.I))

    for match in re.finditer(r"""(?:src|href)=["']([^"']+)["']""", text, flags=re.I):
        candidates.append(match.group(1))
    for match in re.finditer(r"""https?://[^"'<> ]+\.(?:jpg|jpeg|png|webp|gif|bmp)(?:\?[^"'<> ]*)?""", text, flags=re.I):
        candidates.append(match.group(0))
    json_url_patterns = [
        r'"(?:downloadUrl|download_url|contentUrl|content_url|thumbnailUrl|thumbnail_url|url)"\s*:\s*"([^"]+)"',
        r"'(?:downloadUrl|download_url|contentUrl|content_url|thumbnailUrl|thumbnail_url|url)'\s*:\s*'([^']+)'",
    ]
    for pattern in json_url_patterns:
        candidates.extend(re.findall(pattern, text, flags=re.I))

    clean = []
    seen = set()
    for candidate in candidates:
        candidate = urllib.parse.urljoin(base_url, candidate.strip())
        lowered = candidate.lower()
        looks_media = (
            is_probably_image_url(candidate)
            or "download" in lowered
            or "thumbnail" in lowered
            or "public" in lowered and ("onedrive" in lowered or "sharepoint" in lowered)
        )
        if not looks_media:
            continue
        if candidate not in seen:
            seen.add(candidate)
            clean.append(candidate)
    return clean


def microsoft_share_content_url(url):
    token = base64.urlsafe_b64encode(url.encode("utf-8")).decode("ascii").rstrip("=")
    return f"https://api.onedrive.com/v1.0/shares/u!{token}/root/content"


def sharepoint_api_quote(value: str) -> str:
    return urllib.parse.quote(str(value or "").replace("'", "''"), safe="/")


def sharepoint_download_url(web_url: str, server_relative_url: str) -> str:
    quoted = urllib.parse.quote(server_relative_url, safe="/")
    return f"{web_url}/_layouts/15/download.aspx?SourceUrl={quoted}"


class UniversalDownloader:
    def __init__(
        self,
        output_dir,
        workers=DEFAULT_WORKERS,
        timeout=DEFAULT_TIMEOUT,
        log=None,
        skip_existing=False,
        file_mode="all",
        duplicate_detection=False,
        auto_subfolder=False,
        max_retries=2,
        provider_limits=None,
        transfer_progress=None,
        task_events=None,
        proxy=None,
        speed_limit_kbps=0,
    ):
        self.output_dir = Path(output_dir)
        self.workers = min(MAX_WORKERS, max(MIN_WORKERS, int(workers)))
        self.timeout = max(MIN_TIMEOUT, int(timeout))
        self.log = log or (lambda text: None)
        self.skip_existing = bool(skip_existing)
        self.file_mode = file_mode if file_mode in {"images", "media", "all"} else "all"
        self.duplicate_detection = bool(duplicate_detection)
        self.auto_subfolder = bool(auto_subfolder)
        self.max_retries = min(MAX_RETRIES_LIMIT, max(0, int(max_retries)))
        self.transfer_progress = transfer_progress
        self.task_events = task_events
        self.stop_event = None
        self.pause_event = None
        self.current_task = threading.local()
        limits = {key: min(value, self.workers) for key, value in PROVIDER_LIMITS.items()}
        limits.update(provider_limits or {})
        self.provider_limits = {
            key: max(1, min(self.workers, int(value)))
            for key, value in limits.items()
        }
        self.provider_semaphores = {
            key: threading.BoundedSemaphore(value)
            for key, value in self.provider_limits.items()
        }
        self.seen_lock = threading.Lock()
        self.seen_keys = set()
        self.report_lock = threading.Lock()
        self.report_rows = []
        self.session = requests.Session()
        self.session.headers.update({"User-Agent": DEFAULT_USER_AGENT})
        # Optional proxy support ({"http": ..., "https": ...}).
        proxy = proxy or {}
        clean_proxy = {}
        for key in ("http", "https"):
            value = str(proxy.get(key, "") or "").strip()
            if value:
                clean_proxy[key] = value
        if clean_proxy:
            self.session.proxies.update(clean_proxy)
        # Optional download speed limit (KB/s). 0 = unlimited.
        self.speed_limit_kbps = max(0, int(speed_limit_kbps or 0))
        self._speed_lock = threading.Lock()
        self._speed_window_start = 0.0
        self._speed_window_bytes = 0

    def run(self, tasks, progress=None, stop_event=None, pause_event=None):
        self.stop_event = stop_event
        self.pause_event = pause_event
        self.output_dir.mkdir(parents=True, exist_ok=True)
        results = []
        total = len(tasks)
        done = 0

        with ThreadPoolExecutor(max_workers=self.workers) as executor:
            pending_tasks = iter(tasks)
            futures = {}

            def submit_next():
                if stop_event is not None and stop_event.is_set():
                    return False
                while pause_event is not None and pause_event.is_set():
                    if stop_event is not None and stop_event.is_set():
                        return False
                    time.sleep(PAUSE_POLL_INTERVAL)
                try:
                    task_item = next(pending_tasks)
                except StopIteration:
                    return False
                futures[executor.submit(self.download_task, task_item)] = task_item
                return True

            for _ in range(self.workers):
                if not submit_next():
                    break

            while futures:
                done_futures, _ = wait(futures, return_when=FIRST_COMPLETED)
                for future in done_futures:
                    task = futures.pop(future)
                    try:
                        result = future.result()
                    except Exception as exc:
                        result = DownloadResult("failed", 0, str(exc))
                    done += 1
                    results.append((task, result))
                    self.append_report(task, result)
                    if progress:
                        progress(done, total, task, result)
                    if stop_event is None or not stop_event.is_set():
                        submit_next()

            if stop_event is not None and stop_event.is_set():
                executor.shutdown(wait=False, cancel_futures=True)

        self.write_xlsx_report()
        return results

    def download_task(self, task):
        self.output_dir.mkdir(parents=True, exist_ok=True)
        target = self.output_dir / sanitize_name(task.folder)
        provider = provider_name(task.url)
        self.log(f"[{task.row}] {task.folder}: {provider} -> {task.url}")

        if self.skip_existing and self.folder_has_files(target):
            return DownloadResult("skipped", 0, "Skipped because output folder already has files.")

        if self.duplicate_detection:
            key = (str(target.resolve()).lower(), task.url.strip().lower())
            with self.seen_lock:
                if key in self.seen_keys:
                    return DownloadResult("skipped_duplicate", 0, "Skipped duplicate URL in same folder.")
                self.seen_keys.add(key)

        temp_target = Path(tempfile.mkdtemp(prefix=f".tmp_{sanitize_name(task.folder)}_", dir=self.output_dir))
        started = time.monotonic()
        result = DownloadResult("failed", 0, "Download did not start.")
        attempts = 0
        try:
            semaphore = self.provider_semaphores.get(provider, self.provider_semaphores["generic"])
            with semaphore:
                for attempts in range(1, self.max_retries + 2):
                    if self.task_events:
                        self.task_events("start", task, provider, attempts, None)
                    event_result = None
                    try:
                        self.current_task.task = task
                        result = self._download_provider(provider, task.url, temp_target)
                        event_result = result
                    finally:
                        self.current_task.task = None
                        if self.task_events:
                            self.task_events("finish", task, provider, attempts, event_result)
                    if result.status == "ok" or not self.is_retryable_result(result):
                        break
                    delay = min(RETRY_BACKOFF_MAX, attempts)
                    self.log(
                        f"[{task.row}] retry {attempts}/{self.max_retries} in {delay}s "
                        f"({result.status})"
                    )
                    if self.stop_event and self.stop_event.is_set():
                        break
                    time.sleep(delay)
            result.attempts = attempts
            result.duration = round(time.monotonic() - started, 2)

            if result.status == "ok" and self.folder_has_files(temp_target):
                target.mkdir(parents=True, exist_ok=True)
                moved = self.move_downloaded_contents(temp_target, target)
                if self.auto_subfolder:
                    self._apply_auto_subfolder(target)
                return DownloadResult(
                    "ok",
                    moved or result.files,
                    result.message,
                    bytes_downloaded=result.bytes_downloaded,
                    attempts=attempts,
                    duration=result.duration,
                )
            return result
        finally:
            shutil.rmtree(temp_target, ignore_errors=True)
            try:
                if target.exists() and not any(target.iterdir()):
                    target.rmdir()
            except Exception:
                pass

    def _download_provider(self, provider, url, target):
        if provider == "sharepoint_onedrive":
            return self.download_sharepoint(url, target)
        if provider == "google_drive":
            return self.download_google_drive(url, target)
        if provider == "dropbox":
            return self.download_dropbox(url, target)
        return self.download_http_or_page(url, target)

    def is_retryable_result(self, result):
        if result.status in RETRYABLE_STATUSES:
            return True
        text = str(result.message or "").lower()
        return result.status == "failed" and any(term in text for term in RETRYABLE_MESSAGE_FRAGMENTS)

    def move_downloaded_contents(self, source, target):
        moved = 0
        source = Path(source)
        target = Path(target)
        for path in source.rglob("*"):
            if not path.is_file():
                continue
            if is_internal_metadata_file(path):
                path.unlink(missing_ok=True)
                continue
            if not self.is_allowed_payload(path.name):
                path.unlink(missing_ok=True)
                continue
            relative = path.relative_to(source)
            dest_dir = target / relative.parent
            dest = unique_path(dest_dir, relative.name)
            shutil.move(str(path), str(dest))
            moved += 1
        return moved

    def _apply_auto_subfolder(self, target_dir):
        target_dir = Path(target_dir)
        if not target_dir.exists() or not target_dir.is_dir():
            return
        for file_path in list(target_dir.iterdir()):
            if not file_path.is_file():
                continue
            parts = re.split(r"[_.\-]", file_path.stem, maxsplit=1)
            if len(parts) <= 1 or not parts[0]:
                continue
            sub_dir = target_dir / sanitize_name(parts[0])
            sub_dir.mkdir(exist_ok=True)
            try:
                shutil.move(str(file_path), str(unique_path(sub_dir, file_path.name)))
            except Exception:
                pass

    def folder_has_files(self, target):
        return Path(target).exists() and any(
            path.is_file() and ".partials" not in path.parts and not is_internal_metadata_file(path)
            for path in Path(target).rglob("*")
        )

    def is_allowed_payload(self, filename, content_type=""):
        if is_internal_metadata_file(filename):
            return False
        if self.file_mode == "all":
            return True
        suffix = Path(filename or "").suffix.lower()
        content_type = (content_type or "").lower()
        is_image = suffix in IMAGE_EXTENSIONS or content_type.startswith("image/")
        is_video = suffix in VIDEO_EXTENSIONS or content_type.startswith("video/")
        if self.file_mode == "images":
            return is_image
        if self.file_mode == "media":
            return is_image or is_video
        return True

    def download_dropbox(self, url, target):
        direct_url = normalize_dropbox(url)
        # Dropbox returns its web page for browser-like user agents on folder links.
        # A downloader-style user agent keeps public folder shares as zip payloads.
        headers = {
            "User-Agent": DROPBOX_USER_AGENT,
            "Accept": "application/zip,application/octet-stream,*/*",
        }
        result = self.download_http(direct_url, target, headers=headers)
        if result.status == "ok":
            return result
        if is_dropbox_folder(url) and "HTML" in result.message:
            return DownloadResult(
                "failed",
                0,
                "Received Dropbox folder page instead of a zip. Make sure the link is public and download-enabled.",
            )
        return result

    def download_sharepoint(self, url, target):
        url_lower = url.lower()
        if "/:i:/" in url_lower or "/:u:/" in url_lower or "/:v:/" in url_lower or "/:b:/" in url_lower:
            return DownloadResult("login_required", 0, "SharePoint direct files require browser fallback logic.")

        parts = urllib.parse.urlsplit(url)
        query = dict(urllib.parse.parse_qsl(parts.query, keep_blank_values=True))
        query["download"] = "1"
        direct_url = urllib.parse.urlunsplit(
            (parts.scheme, parts.netloc, parts.path, urllib.parse.urlencode(query), parts.fragment)
        )
        candidates = [direct_url, microsoft_share_content_url(url)]
        last_result = DownloadResult("failed", 0, "SharePoint/OneDrive direct download failed.")
        for candidate in candidates:
            direct_result = self.download_http(candidate, target, allow_html=True)
            if direct_result.status == "ok":
                return direct_result
            last_result = direct_result

        folder_result = self.download_sharepoint_folder_api(url, target)
        if folder_result.status == "ok":
            return folder_result

        if not is_sharepoint_folder_share(url):
            parsed_result = self.download_onedrive_viewer_media(url, target)
            if parsed_result.status == "ok":
                return parsed_result

        if not SHAREPOINT_CORE:
            return folder_result if folder_result.status != "failed" else DownloadResult(
                "missing_core",
                0,
                "SharePoint folder could not be listed and legacy core module is missing.",
            )
        # Folder shares and protected links still need the specialized core.
        result = SHAREPOINT_CORE.download_url(url, target, mode="folder")
        if result.status == "ok":
            return DownloadResult(result.status, result.files, result.message)
        if "login" in str(result.message).lower() or "access" in str(result.message).lower():
            return DownloadResult(
                "login_required",
                0,
                "Protected OneDrive/SharePoint link. Use Auth Session once, then enable Browser fallback.",
            )
        return DownloadResult(result.status, result.files, result.message)

    def resolve_sharepoint_folder(self, url):
        response = self.session.get(url, timeout=self.timeout, allow_redirects=True)
        response.raise_for_status()
        final_url = response.url
        parts = urllib.parse.urlsplit(final_url)
        query = urllib.parse.parse_qs(parts.query)
        folder_path = query.get("id", [""])[0]
        if not folder_path:
            return "", "", final_url
        web_url = final_url.split("/_layouts/")[0] if "/_layouts/" in final_url else f"{parts.scheme}://{parts.netloc}"
        return web_url, urllib.parse.unquote(folder_path), final_url

    def sharepoint_folder_items(self, web_url, folder_path):
        quoted = sharepoint_api_quote(folder_path)
        endpoint = (
            f"{web_url}/_api/web/GetFolderByServerRelativeUrl('{quoted}')"
            "?$expand=Files,Folders"
        )
        response = self.session.get(
            endpoint,
            timeout=self.timeout,
            headers={"Accept": "application/json;odata=nometadata"},
        )
        response.raise_for_status()
        data = response.json()
        return data.get("Files", []) or [], data.get("Folders", []) or []

    def download_sharepoint_folder_api(self, url, target):
        try:
            web_url, folder_path, final_url = self.resolve_sharepoint_folder(url)
        except requests.Timeout as exc:
            return DownloadResult("timeout", 0, f"SharePoint folder resolve timed out: {exc}")
        except requests.ConnectionError as exc:
            return DownloadResult("network_error", 0, f"SharePoint folder resolve network error: {exc}")
        except Exception as exc:
            return DownloadResult("failed", 0, f"SharePoint folder resolve failed: {exc}")

        if not web_url or not folder_path:
            return DownloadResult("no_files", 0, "SharePoint folder path was not exposed by the shared link.")

        saved = 0
        bytes_downloaded = 0
        visited = set()
        messages = []

        def walk(current_folder, relative_parts):
            nonlocal saved, bytes_downloaded
            if current_folder in visited:
                return
            visited.add(current_folder)
            files, folders = self.sharepoint_folder_items(web_url, current_folder)
            local_target = Path(target).joinpath(*relative_parts) if relative_parts else Path(target)
            for file_item in files:
                name = file_item.get("Name") or Path(file_item.get("ServerRelativeUrl", "")).name
                server_relative = file_item.get("ServerRelativeUrl")
                if not server_relative:
                    continue
                if not self.is_allowed_payload(name):
                    continue
                file_url = sharepoint_download_url(web_url, server_relative)
                result = self.download_http(file_url, local_target, fallback_name=name)
                if result.status == "ok":
                    saved += result.files
                    bytes_downloaded += result.bytes_downloaded
                else:
                    messages.append(f"{name}: {result.message}")
            for folder_item in folders:
                folder_name = folder_item.get("Name") or Path(folder_item.get("ServerRelativeUrl", "")).name
                if not folder_name or folder_name.lower() == "forms":
                    continue
                child_relative = folder_item.get("ServerRelativeUrl")
                if child_relative:
                    walk(child_relative, [*relative_parts, sanitize_name(folder_name)])

        try:
            walk(folder_path, [])
        except requests.Timeout as exc:
            return DownloadResult("timeout", saved, f"SharePoint folder API timed out after {saved} file(s): {exc}")
        except requests.ConnectionError as exc:
            return DownloadResult("network_error", saved, f"SharePoint folder API network error after {saved} file(s): {exc}")
        except Exception as exc:
            if saved:
                return DownloadResult("ok", saved, f"Saved {saved} file(s); some SharePoint items failed: {exc}", bytes_downloaded=bytes_downloaded)
            return DownloadResult("failed", 0, f"SharePoint folder API failed: {exc}")

        if saved:
            return DownloadResult("ok", saved, f"Saved {saved} file(s) from SharePoint folder", bytes_downloaded=bytes_downloaded)
        return DownloadResult("no_files", 0, "; ".join(messages[:2]) or "SharePoint folder contained no downloadable files.")

    def download_onedrive_viewer_media(self, url, target):
        try:
            response = self.session.get(url, timeout=self.timeout)
            response.raise_for_status()
            links = image_links_from_html(response.text, response.url)
        except requests.Timeout as exc:
            return DownloadResult("timeout", 0, f"Viewer parse timed out: {exc}")
        except requests.ConnectionError as exc:
            return DownloadResult("network_error", 0, f"Viewer parse network error: {exc}")
        except Exception as exc:
            return DownloadResult("failed", 0, f"Viewer parse failed: {exc}")

        if not links:
            return DownloadResult("no_images", 0, "OneDrive viewer page parsed, but image URLs were not exposed.")

        saved = 0
        messages = []
        for index, image_url in enumerate(links[:MAX_ONEDRIVE_IMAGES], start=1):
            result = self.download_http(image_url, target, fallback_name=f"onedrive_image_{index}.jpg")
            if result.status == "ok":
                saved += result.files
            else:
                messages.append(result.message)
        if saved:
            return DownloadResult("ok", saved, f"Saved {saved} image(s) from OneDrive viewer page")
        return DownloadResult("failed", 0, "; ".join(messages[:2]) or "OneDrive viewer images could not be downloaded.")

    def download_google_drive(self, url, target):
        direct_result = self.download_google_drive_direct(url, target)
        if direct_result.status == "ok":
            return direct_result

        if gdown is None:
            file_id = google_drive_id(url)
            if file_id and not google_drive_is_folder(url):
                return direct_result
            return DownloadResult(
                "needs_gdown",
                0,
                "Google Drive folders require gdown. Run: pip install -r requirements.txt",
            )

        before = self.count_files(target)
        before_bytes = self.count_bytes(target)
        try:
            if google_drive_is_folder(url):
                gdown.download_folder(url=url, output=str(target), quiet=True, use_cookies=False)
            else:
                file_id = google_drive_id(url)
                output = str(target / "gdrive_download")
                gdown.download(id=file_id, url=None if file_id else url, output=output, quiet=True, use_cookies=False)
                self.fix_extension_after_download(Path(output))
            self.cleanup_internal_metadata(target)
            after = self.count_files(target)
            after_bytes = self.count_bytes(target)
            saved = max(0, after - before)
            if saved:
                return DownloadResult(
                    "ok",
                    saved,
                    f"Saved {saved} file(s) from Google Drive",
                    bytes_downloaded=max(0, after_bytes - before_bytes),
                )
            return DownloadResult("failed", 0, "Google Drive returned no downloadable file.")
        except Exception as exc:
            if direct_result.status != "failed":
                return direct_result
            return DownloadResult("failed", 0, str(exc))

    def download_google_drive_direct(self, url, target):
        file_id = google_drive_id(url)
        if not file_id or google_drive_is_folder(url):
            return DownloadResult("failed", 0, "Google Drive direct file id not found.")
        candidates = [
            f"https://drive.usercontent.google.com/download?id={file_id}&export=download",
            f"https://drive.google.com/uc?export=download&id={file_id}",
        ]
        last = DownloadResult("failed", 0, "Google Drive direct download failed.")
        for candidate in candidates:
            result = self.download_http(candidate, target, allow_html=True)
            if result.status == "ok":
                return result
            last = result
        return last

    def fix_extension_after_download(self, path):
        if not path.exists() or path.suffix:
            return
        with path.open("rb") as fh:
            head = fh.read(MAGIC_BYTE_BUFFER)
        ext = ".bin"
        if head.startswith(b"\xff\xd8"):
            ext = ".jpg"
        elif head.startswith(b"\x89PNG"):
            ext = ".png"
        elif head.startswith(b"PK"):
            ext = ".zip"
        new_path = unique_path(path.parent, path.name + ext)
        path.replace(new_path)
        if ext == ".zip":
            self.extract_archive(new_path, path.parent)

    def download_http_or_page(self, url, target):
        result = self.download_http(url, target, allow_html=True)
        if result.status != "html_page":
            return result

        try:
            response = self.session.get(url, timeout=self.timeout)
            response.raise_for_status()
            links = image_links_from_html(response.text, response.url)
        except Exception as exc:
            return DownloadResult("failed", 0, f"Page parse failed: {exc}")

        if not links:
            return DownloadResult("no_images", 0, "Page opened, but no image links were found.")

        saved = 0
        messages = []
        for index, image_url in enumerate(links[:MAX_PAGE_IMAGES], start=1):
            item_result = self.download_http(image_url, target, fallback_name=f"image_{index}.jpg")
            if item_result.status == "ok":
                saved += item_result.files
            else:
                messages.append(item_result.message)
        if saved:
            return DownloadResult("ok", saved, f"Saved {saved} image(s) parsed from page")
        return DownloadResult("failed", 0, "; ".join(messages[:2]) or "Images could not be downloaded.")

    def download_http(self, url, target, allow_html=False, fallback_name="download", headers=None):
        response = None
        partial_dir = self.output_dir / ".partials"
        partial_dir.mkdir(parents=True, exist_ok=True)
        partial_path = partial_dir / f"{hashlib.sha256(url.encode('utf-8')).hexdigest()}.part"
        existing_size = partial_path.stat().st_size if partial_path.exists() else 0
        request_headers = dict(headers or {})
        if existing_size:
            request_headers["Range"] = f"bytes={existing_size}-"
        try:
            client = requests if headers else self.session
            response = client.get(
                url,
                stream=True,
                timeout=self.timeout,
                allow_redirects=True,
                headers=request_headers or None,
            )
            if response.status_code in (403, 401) and not headers:
                # Akamai/Cloudflare often block spoofed browser UAs due to TLS fingerprint mismatch.
                # Retry with the honest python-requests UA.
                response.close()
                request_headers["User-Agent"] = requests.utils.default_user_agent()
                client = requests
                response = client.get(
                    url,
                    stream=True,
                    timeout=self.timeout,
                    allow_redirects=True,
                    headers=request_headers,
                )
            if response.status_code == 416 and existing_size:
                partial_path.unlink(missing_ok=True)
                return DownloadResult("network_error", 0, "Server rejected the partial-file resume range.")
            if response.status_code == 429:
                return DownloadResult("rate_limited", 0, "HTTP 429 rate limited.")
            if response.status_code >= 500:
                return DownloadResult("server_error", 0, f"HTTP {response.status_code} server error.")
            response.raise_for_status()
            content_type = response.headers.get("content-type", "")
            if "text/html" in content_type.lower():
                if allow_html:
                    return DownloadResult("html_page", 0, "HTML page received")
                return DownloadResult("failed", 0, "Received an HTML/login page instead of a direct file.")

            resumed = existing_size > 0 and response.status_code == 206
            if existing_size and not resumed:
                existing_size = 0
            mode = "ab" if resumed else "wb"
            bytes_received = 0
            with partial_path.open(mode) as tmp:
                for chunk in response.iter_content(CHUNK_SIZE):
                    if self.stop_event and self.stop_event.is_set():
                        break
                    if chunk:
                        tmp.write(chunk)
                        bytes_received += len(chunk)
                        self._throttle(len(chunk))
                        if self.transfer_progress:
                            task = getattr(self.current_task, "task", None)
                            try:
                                self.transfer_progress(len(chunk), url, task)
                            except TypeError:
                                self.transfer_progress(len(chunk), url)

            if self.stop_event and self.stop_event.is_set():
                partial_path.unlink(missing_ok=True)
                return DownloadResult("failed", 0, "Download cancelled by user.")

            expected = int(response.headers.get("content-length") or 0)
            if expected and bytes_received < expected:
                return DownloadResult(
                    "network_error",
                    0,
                    f"Incomplete download: received {bytes_received} of {expected} bytes; partial saved.",
                    bytes_downloaded=bytes_received,
                )

            validation_error = self.validate_downloaded_file(partial_path, content_type)
            if validation_error:
                partial_path.unlink(missing_ok=True)
                return DownloadResult("corrupt_file", 0, validation_error, bytes_downloaded=bytes_received)

            if zipfile.is_zipfile(partial_path):
                count = self.extract_archive(partial_path, target)
                partial_path.unlink(missing_ok=True)
                if count:
                    return DownloadResult(
                        "ok",
                        count,
                        f"Extracted {count} file(s) from zip",
                        bytes_downloaded=bytes_received,
                    )
                return DownloadResult("no_files", 0, "Zip was downloaded, but no usable files were found inside.")

            name = filename_from_headers(response.headers) or filename_from_url(response.url, content_type)
            if name == "download.bin" and fallback_name:
                name = fallback_name
            if not self.is_allowed_payload(name, content_type):
                partial_path.unlink(missing_ok=True)
                return DownloadResult("skipped_type", 0, f"Skipped by file mode: {name}")
            dest = unique_path(target, name)
            shutil.move(str(partial_path), dest)
            resume_note = " (resumed)" if resumed else ""
            return DownloadResult(
                "ok",
                1,
                f"Saved {dest.name}{resume_note}",
                bytes_downloaded=bytes_received,
            )
        except requests.Timeout as exc:
            return DownloadResult("timeout", 0, f"Request timed out: {exc}")
        except requests.ConnectionError as exc:
            return DownloadResult("network_error", 0, f"Network error: {exc}")
        except requests.HTTPError as exc:
            code = exc.response.status_code if exc.response is not None else 0
            return DownloadResult("http_error", 0, f"HTTP {code}: {exc}")
        except Exception as exc:
            return DownloadResult("failed", 0, str(exc))
        finally:
            if response is not None:
                response.close()

    def _throttle(self, byte_count: int) -> None:
        """Enforce an optional global download speed limit (token-bucket).

        When ``speed_limit_kbps`` is 0 this is a no-op. Otherwise we accumulate
        downloaded bytes in a 1-second sliding window and sleep if the window
        exceeds the configured KB/s budget. Thread-safe so all worker threads
        share the same budget.
        """
        if not self.speed_limit_kbps:
            return
        budget = self.speed_limit_kbps * 1024  # bytes per second
        with self._speed_lock:
            now = time.monotonic()
            if self._speed_window_start == 0.0:
                self._speed_window_start = now
            self._speed_window_bytes += int(byte_count)
            elapsed = now - self._speed_window_start
            if elapsed >= 1.0:
                # Reset window each second.
                self._speed_window_start = now
                self._speed_window_bytes = 0
            elif self._speed_window_bytes >= budget:
                overshoot = (self._speed_window_bytes - budget) / max(budget, 1)
                time.sleep(min(1.0, overshoot))
                self._speed_window_start = time.monotonic()
                self._speed_window_bytes = 0

    def validate_downloaded_file(self, path, content_type=""):
        path = Path(path)
        if not path.exists() or path.stat().st_size == 0:
            return "Downloaded file is empty."
        with path.open("rb") as fh:
            head = fh.read(VALIDATION_BUFFER)
        lowered = head.lstrip().lower()
        if lowered.startswith((b"<!doctype html", b"<html", b"<head", b"<body")):
            return "Server returned an HTML/login page instead of a file."
        content_type = (content_type or "").lower()
        if content_type.startswith("image/"):
            signatures = (
                head.startswith(b"\xff\xd8\xff"),
                head.startswith(b"\x89PNG\r\n\x1a\n"),
                head.startswith((b"GIF87a", b"GIF89a")),
                head.startswith(b"BM"),
                head.startswith((b"II*\x00", b"MM\x00*")),
                len(head) >= 12 and head[8:12] == b"WEBP",
                b"ftypheic" in head[:32] or b"ftypheif" in head[:32],
            )
            if not any(signatures):
                return f"Image payload failed signature validation ({content_type})."
        return ""

    def extract_archive(self, archive_path, target):
        count = 0
        target = Path(target)
        with zipfile.ZipFile(archive_path) as zf:
            for item in zf.infolist():
                if item.is_dir():
                    continue
                source_name = Path(item.filename)
                if source_name.name.startswith("."):
                    continue
                if not self.is_allowed_payload(source_name.name):
                    continue
                dest = unique_path(target, source_name.name)
                with zf.open(item) as src, dest.open("wb") as dst:
                    shutil.copyfileobj(src, dst)
                count += 1
        return count

    def count_files(self, folder):
        if not Path(folder).exists():
            return 0
        return sum(
            1
            for path in Path(folder).rglob("*")
            if path.is_file() and ".partials" not in path.parts and not is_internal_metadata_file(path)
        )

    def count_bytes(self, folder):
        if not Path(folder).exists():
            return 0
        return sum(
            path.stat().st_size
            for path in Path(folder).rglob("*")
            if path.is_file() and ".partials" not in path.parts and not is_internal_metadata_file(path)
        )

    def cleanup_internal_metadata(self, folder):
        if not Path(folder).exists():
            return 0
        removed = 0
        for path in Path(folder).rglob("*"):
            if path.is_file() and is_internal_metadata_file(path):
                path.unlink(missing_ok=True)
                removed += 1
        return removed

    def append_report(self, task, result):
        row = [
            time.strftime("%Y-%m-%d %H:%M:%S"),
            task.row,
            task.folder,
            task.url,
            provider_name(task.url),
            result.status,
            result.files,
            result.message,
            result.attempts,
            result.duration,
            result.bytes_downloaded,
            str((self.output_dir / sanitize_name(task.folder)).resolve()),
        ]
        with self.report_lock:
            self.report_rows.append(row)

    def write_xlsx_report(self):
        if openpyxl is None or not self.report_rows:
            return
        xlsx_path = self.output_dir / LOG_NAME
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Download Report"
        headers = [
            "time", "row", "folder", "url", "provider", "status", "files",
            "message", "attempts", "duration_seconds", "bytes_downloaded", "output_folder"
        ]
        sheet.append(headers)
        for row in self.report_rows:
            sheet.append(row)
        for column_cells in sheet.columns:
            width = min(max(len(str(cell.value or "")) for cell in column_cells) + 2, 70)
            sheet.column_dimensions[column_cells[0].column_letter].width = width
        workbook.save(xlsx_path)
        self.report_rows.clear()


def filename_from_headers(headers):
    disposition = headers.get("content-disposition", "")
    match = re.search(r"filename\*=UTF-8''([^;]+)", disposition, flags=re.I)
    if match:
        return sanitize_name(urllib.parse.unquote(match.group(1)))
    match = re.search(r'filename="?([^";]+)"?', disposition, flags=re.I)
    if match:
        return sanitize_name(match.group(1))
    return ""


def rows_from_pasted_text(text):
    text = (text or "").strip()
    if not text:
        return []
    lines = text.splitlines()
    delimiter = "\t" if any("\t" in line for line in lines) else ","
    if delimiter == "," and not any("," in line for line in lines):
        return [[f"row_{idx}", line.strip()] for idx, line in enumerate(lines, start=1) if line.strip()]
    return [[cell.strip() for cell in row] for row in csv.reader(lines, delimiter=delimiter) if any(cell.strip() for cell in row)]


def rows_from_file(path):
    path = Path(path)
    if path.suffix.lower() == ".csv":
        with path.open(newline="", encoding="utf-8-sig") as fh:
            return [[cell.strip() for cell in row] for row in csv.reader(fh) if any(cell.strip() for cell in row)]
    if path.suffix.lower() in {".xlsx", ".xlsm"}:
        if openpyxl is None:
            raise RuntimeError("openpyxl is missing. Run: pip install -r requirements.txt")
        workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
        sheet = workbook.active
        rows = []
        for row in sheet.iter_rows(values_only=True):
            values = ["" if value is None else str(value).strip() for value in row]
            if any(values):
                rows.append(values)
        return rows
    raise RuntimeError("Only CSV/XLSX supported.")


def tasks_from_rows(rows):
    tasks = []
    start_index = 1
    headers = []
    if rows:
        first = [str(cell).lower() for cell in rows[0]]
        if not any(is_url(cell) for cell in first):
            start_index = 2
            headers = [str(cell).strip() for cell in rows[0]]
            rows = rows[1:]

    for row_number, row in enumerate(rows, start=start_index):
        clean = [str(cell).strip() for cell in row if str(cell).strip()]
        if not clean:
            continue
        urls = [cell for cell in clean if is_url(cell)]
        if not urls:
            continue
        row_data = dict(zip(headers, row)) if headers else {}
        folder_candidates = [cell for cell in clean if not is_url(cell)]
        folder = folder_candidates[0] if folder_candidates else f"row_{row_number}"
        for url in urls:
            tasks.append(LinkTask(row_number, folder, url, row_data))
    return tasks


if TKINTER_AVAILABLE:
  class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title(f"{APP_TITLE} v{APP_VERSION}")
        self.geometry("1120x760")
        self.minsize(980, 680)
        self.configure(bg="#111827")
        self.queue = queue.Queue()
        self.loaded_rows = None
        self.worker = None
        self.output_var = tk.StringVar(value=str((Path.cwd() / "downloads").resolve()))
        self.workers_var = tk.IntVar(value=50)
        self.timeout_var = tk.IntVar(value=60)
        self.status_var = tk.StringVar(value="Ready")
        self.summary_var = tk.StringVar(value="0 links | 0 ok | 0 failed | 0 files")
        self._build_ui()
        self.after(100, self._drain_queue)

    def _build_ui(self):
        style = ttk.Style(self)
        try:
            style.theme_use("clam")
        except Exception:
            pass
        style.configure(".", font=("Segoe UI", 10), background="#111827")
        style.configure("Root.TFrame", background="#111827")
        style.configure("Panel.TFrame", background="#1f2937")
        style.configure("Title.TLabel", background="#111827", foreground="#f9fafb", font=("Segoe UI Semibold", 22))
        style.configure("Sub.TLabel", background="#111827", foreground="#9ca3af")
        style.configure("Panel.TLabel", background="#1f2937", foreground="#f9fafb")
        style.configure("Muted.Panel.TLabel", background="#1f2937", foreground="#9ca3af")
        style.configure("Stat.TLabel", background="#1f2937", foreground="#93c5fd", font=("Segoe UI Semibold", 12))
        style.configure("TButton", padding=(12, 8))
        style.configure("Accent.TButton", padding=(18, 10), font=("Segoe UI Semibold", 10))
        style.configure("TEntry", padding=(8, 6))
        style.configure("Horizontal.TProgressbar", thickness=13)

        outer = ttk.Frame(self, padding=18, style="Root.TFrame")
        outer.pack(fill="both", expand=True)
        outer.columnconfigure(0, weight=1)
        outer.rowconfigure(3, weight=1)

        header = ttk.Frame(outer, style="Root.TFrame")
        header.grid(row=0, column=0, sticky="ew", pady=(0, 16))
        header.columnconfigure(0, weight=1)
        ttk.Label(header, text="Media Download Manager", style="Title.TLabel").grid(row=0, column=0, sticky="w")
        ttk.Label(
            header,
            text="Saves mixed image links and public folder shares into barcode-named folders.",
            style="Sub.TLabel",
        ).grid(row=1, column=0, sticky="w", pady=(3, 0))
        ttk.Label(header, textvariable=self.summary_var, style="Stat.TLabel").grid(row=0, column=1, rowspan=2, sticky="e")

        top = tk.Frame(outer, bg="#1f2937", padx=14, pady=14, highlightthickness=1, highlightbackground="#374151")
        top.grid(row=1, column=0, sticky="ew")
        top.columnconfigure(1, weight=1)
        ttk.Label(top, text="Output Folder", style="Panel.TLabel").grid(row=0, column=0, sticky="w")
        ttk.Entry(top, textvariable=self.output_var).grid(row=0, column=1, sticky="ew", padx=10)
        ttk.Button(top, text="Browse", command=self.choose_output).grid(row=0, column=2)
        ttk.Button(top, text="Open", command=self.open_output).grid(row=0, column=3, padx=(8, 0))

        options = tk.Frame(outer, bg="#1f2937", padx=14, pady=12, highlightthickness=1, highlightbackground="#374151")
        options.grid(row=2, column=0, sticky="ew", pady=(10, 10))
        ttk.Button(options, text="Load Excel/CSV", command=self.load_file).pack(side="left")
        ttk.Button(options, text="Clear", command=self.clear_input).pack(side="left", padx=(8, 18))
        ttk.Label(options, text="Workers", style="Panel.TLabel").pack(side="left")
        ttk.Spinbox(options, from_=1, to=50, textvariable=self.workers_var, width=5).pack(side="left", padx=(6, 14))
        ttk.Label(options, text="Timeout", style="Panel.TLabel").pack(side="left")
        ttk.Spinbox(options, from_=10, to=180, textvariable=self.timeout_var, width=5).pack(side="left", padx=(6, 14))
        ttk.Button(options, text="Start Download", style="Accent.TButton", command=self.start).pack(side="left")
        ttk.Label(
            options,
            text="Supported: public image/file links, public Dropbox folders, Google Drive via gdown, SharePoint/OneDrive public links.",
            style="Muted.Panel.TLabel",
        ).pack(side="left", padx=(18, 0))

        body = ttk.PanedWindow(outer, orient="vertical")
        body.grid(row=3, column=0, sticky="nsew")

        input_frame = ttk.LabelFrame(body, text="Input: first column folder/barcode, next columns links")
        input_frame.rowconfigure(0, weight=1)
        input_frame.columnconfigure(0, weight=1)
        self.input_text = ScrolledText(input_frame, height=11, wrap="none")
        self.input_text.configure(font=("Consolas", 10), bg="#f9fafb", fg="#111827", insertbackground="#111827")
        self.input_text.grid(row=0, column=0, sticky="nsew", padx=8, pady=8)
        body.add(input_frame, weight=2)

        log_frame = ttk.LabelFrame(body, text="Progress")
        log_frame.rowconfigure(1, weight=1)
        log_frame.columnconfigure(0, weight=1)
        self.progress = ttk.Progressbar(log_frame, mode="determinate")
        self.progress.grid(row=0, column=0, sticky="ew", padx=8, pady=(8, 4))
        self.log_text = ScrolledText(log_frame, height=13, wrap="word", state="disabled")
        self.log_text.configure(font=("Consolas", 10), bg="#0b1220", fg="#dbeafe", insertbackground="#dbeafe")
        self.log_text.grid(row=1, column=0, sticky="nsew", padx=8, pady=(4, 8))
        body.add(log_frame, weight=3)

        status = ttk.Label(outer, textvariable=self.status_var)
        status.grid(row=4, column=0, sticky="w", pady=(8, 0))

    def choose_output(self):
        path = filedialog.askdirectory(title="Select output folder")
        if path:
            self.output_var.set(path)

    def open_output(self):
        path = Path(self.output_var.get()).expanduser()
        path.mkdir(parents=True, exist_ok=True)
        open_in_file_manager(path.resolve())

    def load_file(self):
        path = filedialog.askopenfilename(
            title="Select Excel/CSV",
            filetypes=[("Excel/CSV", "*.xlsx *.xlsm *.csv"), ("All files", "*.*")]
        )
        if not path:
            return
        try:
            self.loaded_rows = rows_from_file(path)
            self.input_text.delete("1.0", "end")
            self.input_text.insert("1.0", "\n".join("\t".join(row) for row in self.loaded_rows[:MAX_DISPLAYED_ROWS]))
            self.write_log(f"Loaded {len(self.loaded_rows)} row(s) from {path}\n")
        except Exception as exc:
            messagebox.showerror(APP_TITLE, str(exc))

    def clear_input(self):
        self.loaded_rows = None
        self.input_text.delete("1.0", "end")

    def start(self):
        if self.worker and self.worker.is_alive():
            return
        rows = self.loaded_rows or rows_from_pasted_text(self.input_text.get("1.0", "end"))
        tasks = tasks_from_rows(rows)
        if not tasks:
            messagebox.showwarning(APP_TITLE, "No valid links found.")
            return
        output = self.output_var.get().strip()
        if not output:
            messagebox.showwarning(APP_TITLE, "Please select an output folder.")
            return

        report = Path(output) / LOG_NAME
        report.unlink(missing_ok=True)
        self.progress.configure(value=0, maximum=len(tasks))
        self.status_var.set(f"Starting {len(tasks)} link(s)...")
        self.summary_var.set(f"{len(tasks)} links | 0 ok | 0 failed | 0 files")
        self.write_log(f"\nStarting {len(tasks)} link(s). Output: {output}\n")
        self.worker = threading.Thread(target=self._run, args=(tasks, output), daemon=True)
        self.worker.start()

    def _run(self, tasks, output):
        def log(text):
            self.queue.put(("log", text + "\n"))

        def progress(done, total, task, result):
            self.queue.put(("progress", done, total))
            self.queue.put(("log", f"[{task.row}] {task.folder} -> {result.status} | {result.files} file(s) | {result.message}\n"))

        downloader = UniversalDownloader(
            output,
            workers=self.workers_var.get(),
            timeout=self.timeout_var.get(),
            log=log,
        )
        results = downloader.run(tasks, progress=progress)
        ok = sum(1 for _, result in results if result.status == "ok")
        files = sum(result.files for _, result in results)
        failed = len(results) - ok
        self.queue.put(("done", ok, failed, files, str(Path(output).resolve())))

    def write_log(self, text):
        self.log_text.configure(state="normal")
        self.log_text.insert("end", text)
        self.log_text.see("end")
        self.log_text.configure(state="disabled")

    def _drain_queue(self):
        try:
            while True:
                event = self.queue.get_nowait()
                if event[0] == "log":
                    self.write_log(event[1])
                elif event[0] == "progress":
                    self.progress.configure(value=event[1], maximum=max(event[2], 1))
                    self.status_var.set(f"Progress: {event[1]}/{event[2]}")
                elif event[0] == "done":
                    ok, failed, files, output = event[1:]
                    self.status_var.set(f"Done: {ok} ok, {failed} failed, {files} file(s)")
                    self.summary_var.set(f"{ok + failed} links | {ok} ok | {failed} failed | {files} files")
                    self.write_log(f"\nDone. OK: {ok}, Failed: {failed}, Files: {files}\nReport: {Path(output) / LOG_NAME}\n")
                    messagebox.showinfo(APP_TITLE, f"Done.\nOK: {ok}\nFailed: {failed}\nFiles: {files}")
        except queue.Empty:
            pass
        self.after(100, self._drain_queue)


def main():
    if "--cli" in sys.argv:
        if len(sys.argv) < 4:
            print("Usage: python universal_link_downloader.py --cli input.xlsx output_folder")
            return 2
        rows = rows_from_file(sys.argv[2])
        tasks = tasks_from_rows(rows)
        downloader = UniversalDownloader(sys.argv[3], log=print)
        results = downloader.run(tasks)
        print(f"Done. Links: {len(results)}")
        return 0
    if "--classic-gui" not in sys.argv:
        import web_app

        web_app.main()
        return 0
    if TKINTER_AVAILABLE:
        app = App()
        app.mainloop()
        return 0
    else:
        print("GUI mode requires tkinter. Use --cli mode on headless servers.")
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
